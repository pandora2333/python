import types  

from openpyxl import load_workbook  
  
from datetime import date  

from openpyxl import Workbook  
from openpyxl.chart import (  
    LineChart,  
    Reference,  
)  
from openpyxl.chart.axis import DateAxis  
  
wbb = Workbook()  
wss = wbb.active
wb = load_workbook(filename=r'sh300year.xlsx')   # 打开22.xlsx从里面读数据  
   
sheets = wb.get_sheet_names()     
sheet0 = sheets[0]  # 第一个表格的名称  ＃其实感觉没什么用，可以直接写worksheet的名字
wo =  wb.get_sheet_by_name('sum') # 获取特定的 worksheet  填写excel表左下角表的名字
columns1 = wo.columns
content1 = []
rank = [0,0,0,0,0]  # 开一个数组，用于标记交易量最大的期货组合，方便后续拼接  
  
num1 = 0  
num2 = 0  
num3 = 0
for aol in columns1:  
    a = [x.value for x in aol]  
    content1.append(a)
ll = []
rows = wo.rows  
for row in rows:  
    line = [col.value for col in row]  
    ll.append(line)  
  
for roww in ll:  
    wss.append(roww)
c1 = LineChart()            #新建一张图  
c1.title = "Line Chart"     #图的标题  
c1.style = 8                #线条的style  
c1.y_axis.title = 'price'   #y坐标的标题  
#c1.x_axis = DateAxis(crossAx=100)  
c1.x_axis.number_format = 'mm-e'  #规定日期格式  这是月,年格式  
c1.x_axis.majorTimeUnit = "Months"  #规定日期间隔 注意days；Months大写  
c1.x_axis.title = "Date"    #x坐标的标题  
c1.y_axis.scaling.min = 2.5 #y坐标的区间  
c1.y_axis.scaling.max = 3.4  
  
data = Reference(wss, min_col=4, min_row=64981, max_col=4, max_row=105455)    #图像的数据 一列是一条线 若此处取两列，会直接画一个两线的图像  
dataa = Reference(wss, min_col=7, min_row=64981, max_col=7, max_row=105455) #这里要画期货和现货，选择写两列数据，分别画两条线  
c1.add_data(data, titles_from_data=True)  
c1.add_data(dataa, titles_from_data=True)  
#s2 = c1.series[0]  
#s1 = c1.series[1]  
dates = Reference(wss, min_col=1, min_row=64981, max_row=105455)  
c1.set_categories(dates)  
  
  
#s2.smooth = True # Make the line smooth  
wss.add_chart(c1, "H1")      # 添加c1这个chart 图的左上角位置在H1处  
  
wbb.save("ratio.xlsx")        # 写的excel文件为line.xlsx

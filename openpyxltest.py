import openpyxl
import datetime
'''
一个excel文件既是一个工作簿，你可以把工作簿看作是一个本子，而本子是由一页一页的纸张装订在一起的。excel中的sheet就是这些纸张，正规的叫法叫做“工作表”。
'''
workspace = openpyxl.Workbook()#建立新的工作簿
ws = workspace.active#工作表(sheet)建立
ws['a1'] = 1
ws.append([2,3,4])
ws.append([5,6,7])
ws['a3']=datetime.datetime.now()
workspace.save('e:/sheet.xlsx')

import openpyxl
wb = openpyxl.load_workbook('e;/sheet.xlsx')#加载excel文件
wb.get_sheet_names()#以列表的方式显示所有的sheet
wb.sheetnames#同上
ws = wb.get_sheet_by_name('sheet')#获取指定工作表,若不存在，则抛出异常
new = wb.create_sheet(index=0,tittle='pandora')#index指明插入的位置，title表名标题，默认是sheet
wb.remove(wb.get_sheet_by_name('sheet'))#必须传入对象，而不能传入：'sheet'
c = ws['a2']
c.row#得到行数
c.column#得到列数
c.coordinate#得到位置：a2
ws['a2'].value#得到指定单元格的值
d = c.offset(02,0)#得到偏移单元格的对象，第一个参数是column，第二个是row
openpyxl.cell.cell.get_column_letter(496)#得到指定列的标识符：’SB‘,excel的进制是26进制
openpyxl.cell.cell.column_index_from_string('sb')#解决了’aaa'问题，将对应表示转化为十进制
for each in ws['a1':'b10']:
    for each_movie in each:
        print(each_movie.value,end=' ')#支持切片操作
for each in ws.rows:
    print(each[0].value)#以行为单位的列表
for each in ws.iter_rows(min_row=2,min_col=1,max_row=4,max_col=2):
    print(each[0].value)#类似于切片
news = wb.copy_worksheet(ws)#复制工作表
#改变标题与颜色
ws1=wb.create_sheet(title='pandora')
ws1.sheet_properties.tabColor='FF0000'#RGB对应，每个一个字节大小
#合并于拆分单元格
ws1 .row_dimensions[2].height =100
ws1.column_dimensions['p'].width=50#行和列的长度单位不同，发现下面比上面长
ws1.merge_cells('a1:c3')#合并单元格a1至c3
ws1.unmerge_cells('a1:c3')#拆分时只能拆分原来合并时的单元格如不能只拆分a1：b2
#冻结窗口
ws1.freeze_panes='b8'#会将上面部分全部锁定
#解冻
ws1.freeze__pans=None#或者=‘a1’
#字体设置
from openpyxl.styles import Font
bold_red_font = Font(bold=True,color='ff0000')#红色加粗字体
ws1.font = bold_red_font#注册字体
italic_strike_blue_16font = Font(size=16,italic =True,strike=True,color='0000ff')#蓝色倾斜中间横杠的16号字体
ws1.font = italic_strike_blue_16font
#纯色填充
from openpyxl.styles import PatternFill
fill = PatternFill(fill_type='solid',fgColor='ffff00')#第一个是参数意义是全填充
ws1.fill = fill#注册
from openpyxl.styles import GradientFill 
fill_2 = GradientFill(fill_type='linear',stop=('ff0000','00ff00'))#第一个是参数意义是线性渐变，颜色范围是ff0000到00ff00
ws1.fill =fill_2
#边框设置
from openpyxl.styles import Border,Side
thin_side = Side(border_style ='thin',color='000000')#白色单线
double_side =Side(border_style='double',color='ff0000')#红色双线
ws1.border = Border(diagonal=thin_side,diagonalUp=True,diagonalDown=True)#边框设置（对角线设置）
ws1.border =Border(left =double_side,top=double_side,right=doeble_side,bottom=double_side)#边框设置之边框线条设置
#文本对齐
from openpyxl.styles import Alignment
center_alignment = Alignment(horizontal='center',vertical='center')#水平对齐，垂直对齐
ws['a1'].alignment = center_alignment
#命名样式
from openpyxl.styles import NamedStyle
hightlight = NamedStyle(name='hightlight')#style风格命名为高亮
hightlight.font = Font(bold= True,size=20)
hightlight.alignment = Alignment(horizontal = 'center',vertical='center')
ws1.add_named_style(hightlight)#注册
ws1['a1'] = 'love'
ws1['b5'].style= hightlight


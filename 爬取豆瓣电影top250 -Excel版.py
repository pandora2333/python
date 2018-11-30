import requests
import bs4
import re
import openpyxl
def open_url(url):
    #使用代理
    #proxies ={...}
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 Edge/16.16299','Host':'movie.douban.com'}
    #res = requests.get(url,headers = headers,proxies=proxies)
    res = requests.get(url,headers=headers)
    #print(res.encoding)
    return res
def find_movies(res):
    soup = bs4.BeautifulSoup(res.text,'html.parser')
    #电影名
    movies = []
    targets = soup.find_all('div',class_ = 'hd')
    for each in targets:
        movies.append(each.a.span.text)
    #评分
    ranks = []
    targets = soup.find_all('span',class_ = 'rating_num')
    for each in targets:
        ranks.append(each.text)
    #资料
    messages = []
    targets = soup.find_all('div',class_='bd')
    for each in targets:
        try:
            messages.append(each.p.text.split('\n')[1].strip()+each.p.text.split('\n')[2].strip())
        except:
            continue
    result = []
    length =  len(movies)
    for i in range(length):
        result.append([movies[i],ranks[i],messages[i]])
    return result
#找出一共有多少个页面
def find_depth(res):
    soup = bs4.BeautifulSoup(res.text,'html.parser')
    depth = soup.find('span',class_='next').previous_sibling.previous_sibling.text
    return int(depth)
def save_to_excel(result):
    table = openpyxl.Workbook()
    ws = table.active
    ws['A1'] = '电影名称'
    ws['B1'] = '评分'
    ws['C1'] = '资料'
    for each in result:
        ws.append(each)
    table.save('e:/豆瓣top250.xlsx')
def main():
    host = 'https://movie.douban.com/top250'#?start=0&filter=
    res = open_url(host)
    depth = find_depth(res)
    result = []
    for i in range(depth):
        url = host+'?start='+str(25*i)
        res = open_url(url)
        result.extend(find_movies(res))
    save_to_excel(result)
if __name__=='__main__':
    main()
    print('爬取完毕！')
            
